import io
import xlsxwriter
from openpyxl import load_workbook
from pathlib import Path


class ExcelWriter:
    """
    Первичный класс для создания Excel документов.
    Данный класс требует наследования для переопределения метода add.
    """

    def_options = {"in_memory": True}

    def _default_file_path(self):
        """
        Возвращает путь по умолчанию к шаблону Excel.
        Здесь нужно указать правильный путь к вашему файлу шаблона.
        """
        return "../data/template_excel.xlsx"

    def __init__(self, file_path=None, options=None):
        """
        Инициализация класса ExcelWriter.
        Загружает рабочую книгу из указанного файла или создает новую, если файл не существует.
        """
        file_path = file_path or self._default_file_path()
        self.wb = load_workbook(file_path) if Path(file_path).exists() else None
        self.output = (
            io.BytesIO()
        )  # Используется для хранения выходного файла в памяти.
        self.workbook = xlsxwriter.Workbook(
            self.output, options=options or self.def_options
        )

    def write(self, new_filename):
        """
        Записывает содержимое в новый файл.
        new_template_excel - имя нового файла, в который будет записано содержимое.
        """
        with self.workbook as workbook:
            if self.wb:  # Если исходная книга была загружена
                self._copy(workbook)  # Копируем данные из исходного файла
            self.configuration(workbook)  # Конфигурируем файл
            self.add(workbook)  # Добавляем дополнительные данные
        self.output.seek(0)  # Сброс указателя в начале выходного потока
        with open(new_filename, "wb") as f:
            f.write(self.output.getvalue())  # Записываем в новый файл

    def _copy(self, workbook):
        """
        Копирует содержимое из исходной книги в новую книгу.
        """
        for sheet_name in self.wb.sheetnames:  # Проходим по всем листам исходной книги
            source_sheet = self.wb[sheet_name]
            worksheet = workbook.add_worksheet(
                sheet_name
            )  # Создаем новый лист в выходной книге

            # Копируем размеры столбцов
            for col_idx, col_dim in source_sheet.column_dimensions.items():
                if col_dim.width is not None:
                    worksheet.set_column(f"{col_idx}:{col_idx}", col_dim.width)

            # Копируем размеры строк
            for row in source_sheet.iter_rows():
                for cell in row:
                    if source_sheet.row_dimensions[cell.row].height:
                        worksheet.set_row(
                            cell.row - 1, source_sheet.row_dimensions[cell.row].height
                        )

            # Копируем данные и стили
            for row_idx, row in enumerate(source_sheet.iter_rows(), start=0):
                for col_idx, cell in enumerate(row, start=0):
                    # Копируем значение
                    worksheet.write(row_idx, col_idx, cell.value)
                    # Применяем шрифт и стили, если есть
                    if cell.has_style:
                        cell_format = workbook.add_format(self._get_cell_format(cell))
                        worksheet.write(row_idx, col_idx, cell.value, cell_format)

    def _get_cell_format(self, cell):
        """
        Получает формат для ячейки.
        Возвращает словарь с параметрами формата для ячейки.
        """
        format_dict = {}
        font = cell.font
        if font.bold:
            format_dict["bold"] = True
        if font.italic:
            format_dict["italic"] = True
        if font.name:
            format_dict["font_name"] = font.name
        if font.size:
            format_dict["font_size"] = font.size
        if font.color and font.color.rgb:  # Проверяем наличие цвета
            format_dict["font_color"] = f"#{font.color.rgb}"  # Используем цвет напрямую

        align = cell.alignment
        if align.horizontal:
            format_dict["align"] = align.horizontal
        if align.vertical:
            format_dict["valign"] = align.vertical

        return format_dict

    def add(self, workbook):
        """
        Метод для добавления данных, требует переопределения в подклассах.
        """
        pass

    def configuration(self, workbook):
        """
        Начальная конфигурация, может быть переопределена в подклассах.
        """
        pass


class ExcelTemplate(ExcelWriter):
    """
    Класс ExcelTemplate для добавления конкретных данных в Excel.
    Наследуется от ExcelWriter.
    """

    def add(self, workbook):
        """
        Добавляет конкретные данные в новый файл Excel.
        Здесь можно определить, какие данные будут добавлены.
        """
        worksheet = workbook.get_worksheet_by_name("Шаблон")  # Получаем нужный лист
        if worksheet:  # Проверяем, существует ли лист
            worksheet.write(
                "G4", "Вставляем данные"
            )  # Записываем пример данных в ячейку A1
            worksheet.write("G6", "Еще данные")  # Пример добавления еще данных
        else:
            print("Лист 'Шаблон' не найден!")

    def configuration(self, workbook):
        """
        Начальная конфигурация документа Excel.
        Здесь можно настроить параметры документа, такие как стили, форматы и т.д.
        """
        worksheet = workbook.get_worksheet_by_name("Шаблон")

        if worksheet:  # Проверяем, существует ли лист
            # Настраиваем ширину нескольких столбцов
            worksheet.set_column("A:A", 30)  # Устанавливаем ширину столбца A
            worksheet.set_column("B:C", 20)  # Устанавливаем ширину столбцов B и C

            # Настраиваем высоту первой строки
            worksheet.set_row(0, 30)  # Устанавливаем высоту первой строки

            # Создаем формат для заголовка
            header_format = workbook.add_format(
                {
                    "bold": True,
                    "font_color": "white",
                    "bg_color": "blue",
                    "align": "center",
                    "valign": "vcenter",
                }
            )

            # Заполняем заголовок
            headers = ["Дата", "Операция", "Сумма"]
            for col, header in enumerate(headers):
                worksheet.write(0, col, header, header_format)

            # Пример формул и начальных данных
            worksheet.write("A3", "2024-01-01")
            worksheet.write("B3", "Продажа")
            worksheet.write("C3", 5000)

            # Вставляем формулу для суммирования в конец столбца C
            worksheet.write_formula(
                "C10", "=SUM(C3:C9)", workbook.add_format({"bold": True})
            )

            # Дополнительные настройки форматов, если необходимо
            currency_format = workbook.add_format({"num_format": "$#,##0"})
            worksheet.set_column(
                "C:C", None, currency_format
            )  # Устанавливаем формат валюты для столбца C
        else:
            print("Лист 'Шаблон' не найден!")


# Пример использования
excel_template = ExcelTemplate()  # Создаем экземпляр класса ExcelTemplate
excel_template.write(
    "../data/newdata/new_template_excel.xlsx"
)  # Записываем содержимое в новый файл
