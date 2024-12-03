from datetime import datetime
import openpyxl
from copy import copy
from openpyxl.styles import Alignment

class ExcelTemplateFiller:
    """Класс для заполнения шаблона Excel данными из словаря, включая динамическое копирование строк."""

    def __init__(self, template_path):
        """
        Инициализация.
        :param template_path: Путь к шаблонному файлу Excel.
        """
        self.template_path = template_path
        self.workbook = None

    def load_template(self):
        """Загружает шаблонный Excel-файл."""
        self.workbook = openpyxl.load_workbook(self.template_path)

    def add_default_values_to_expenses(self, expenses):
        """
        Добавляет значения по умолчанию для каждого расхода.
        :param expenses: Словарь с расходами.
        :return: Обновленный словарь расходов.
        """
        for expense in expenses.values():
            expense['currency'] = 'руб'  # Добавляем валюту
            expense['pay_method'] = 'б\н'  # Добавляем метод оплаты
        return expenses

    def extract_date_components(self, date_str):
        """
        Извлекает день, месяц и год из строки даты в формате "DD.MM.YYYY".
        :param date_str: Дата в виде строки.
        :return: Кортеж (день, месяц, год).
        """
        try:
            date_obj = datetime.strptime(date_str, "%d.%m.%Y")
            return date_obj.day, date_obj.month, date_obj.year
        except ValueError:
            raise ValueError(f"Некорректный формат даты: {date_str}")

    def fill_template(self, data):
        """
        Заполняет шаблон значениями из словаря.
        :param data: Словарь с данными для замены.
        """
        if not self.workbook:
            raise ValueError("Шаблон не загружен. Используйте метод load_template().")

        # Извлекаем компоненты даты и добавляем их в словарь
        day, month, year = self.extract_date_components(data['date'])
        data['day'] = day
        data['month'] = month
        data['year'] = year

        # Добавляем значения по умолчанию в расходы
        data['expenses'] = self.add_default_values_to_expenses(data.get("expenses", {}))

        for sheet in self.workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        # Если это маркер для вложенного словаря {{expenses}}
                        if "{{expenses}}" in cell.value:
                            self._fill_expenses(sheet, cell, data.get("expenses", {}))
                            cell.value = None  # Очищаем ячейку с маркером
                        else:
                            # Заменяем обычные ключи {{key}}
                            for key, value in data.items():
                                placeholder = f"{{{{{key}}}}}"  # Формируем шаблонную метку {{key}}
                                if placeholder in cell.value:
                                    cell.value = cell.value.replace(placeholder, str(value))

    def _fill_expenses(self, sheet, start_cell, expenses):
        """
        Обрабатывает вложенный словарь расходов, копирует строку и добавляет данные.
        :param sheet: Текущий лист Excel.
        :param start_cell: Ячейка, содержащая маркер {{expenses}}.
        :param expenses: Словарь с данными о расходах.
        """
        start_row = start_cell.row  # Начальная строка, где находится маркер
        col_mapping = {cell.value: cell.column for cell in sheet[start_row] if cell.value}  # Карта столбцов

        # Удаляем маркер {{expenses}} и заменяем на данные из словаря
        sheet.delete_rows(start_row)

        # Заполняем данные из expenses
        for idx, (number, expense) in enumerate(expenses.items(), start=0):
            sheet.insert_rows(start_row + idx)  # Вставляем новую строку

            new_row = sheet[start_row + idx]

            # Копируем стиль ячеек из строки с маркером
            for src_cell, dest_cell in zip(sheet[start_row - 1], new_row):
                dest_cell.font = copy(src_cell.font)
                dest_cell.border = copy(src_cell.border)
                dest_cell.fill = copy(src_cell.fill)
                dest_cell.alignment = Alignment(horizontal="left", vertical="center")  # Выровнять по левому краю

            # Заполняем ячейки новой строки данными
            for key, value in expense.items():
                placeholder = f"{{{{{key}}}}}"
                if placeholder in col_mapping:
                    col_idx = col_mapping[placeholder]
                    value_str = str(value)  # Преобразуем данные в строку
                    cell = sheet.cell(row=start_row + idx, column=col_idx, value=value_str)

            # Добавляем нумерацию в первый столбец (номер расхода)
            sheet.cell(row=start_row + idx, column=start_cell.column, value=str(number))

    def save(self, output_path):
        """
        Сохраняет заполненный файл.
        :param output_path: Путь для сохранения нового Excel-файла.
        """
        if not self.workbook:
            raise ValueError("Шаблон не загружен. Используйте метод load_template().")

        self.workbook.save(output_path)


# Пример использования
if __name__ == "__main__":
    # Путь к шаблону и выходному файлу
    template_path = "../data/advance_template.xlsx"
    output_path = "../data/newdata/filled_template.xlsx"

    # Пример словаря с данными
    data = {
        'start': '2024-11-02',
        'stop': '2024-11-10',
        'duration': 9,
        'expenses': {
            1: {'name': 'Суточные, дней: 9', 'date': '2024-11-02 - 2024-11-10', 'prise': 9450},
            2: {'name': 'Москва - Санкт-Петербург', 'date': '2024-11-02', 'prise': 3500, 'comment': 'Скоростной поезд',
                'type': 'Билет'},
            3: {'name': 'Санкт-Петербург - Москва', 'date': '2024-11-10', 'prise': 3500, 'comment': 'Скоростной поезд',
                'type': 'Билет'},
            4: {'name': 'Такси: офис - вокзал', 'date': '2024-11-02', 'prise': 500, 'comment': 'Утренний трансфер',
                'type': 'чек'},
            5: {'name': 'Такси: вокзал - гостиница', 'date': '2024-11-02', 'prise': 700, 'comment': 'Городская поездка',
                'type': 'чек'},
        },
        'receive': 9450,
        'total_received': 17650,
        'total_spent': 17650,
        'balance': 0,
        'place': 'Россия, Москва, Отдел разработки ООО "Компания"',
        'assignment': 'Командировка на проект',
        'status': 'Завершено',
        'code': 'PRJ-12345',
        'user': 'Иванов Иван Иванович',
        'personal_number': '123456',
        'position': 'Ведущий инженер',
        'division': 'ОТК',
        'date': '13.11.2024'  # Дата для маркеров
    }

    # Создаём экземпляр класса
    filler = ExcelTemplateFiller(template_path)

    # Загружаем шаблон
    filler.load_template()

    # Заполняем шаблон данными
    filler.fill_template(data)

    # Сохраняем заполненный файл
    filler.save(output_path)
