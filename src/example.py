import openpyxl
import shutil
import os
from openpyxl.styles import Alignment, Font
from datetime import datetime


class ExcelCopy:
    """
    Класс для работы с Excel файлами.
    """

    def __init__(self, template_path, output_path):
        self.template_path = template_path
        self.output_path = output_path

    def create_copy(self):
        """Создает физическую копию шаблона Excel."""
        if os.path.exists(self.template_path):
            shutil.copy(self.template_path, self.output_path)
        else:
            raise FileNotFoundError(f"Файл не найден: {self.template_path}")

    def update_excel(self, updates):
        """Обновляет данные в существующем Excel файле."""
        try:
            wb = openpyxl.load_workbook(self.output_path)
            ws = wb.active

            # Вносим изменения в указанные ячейки
            for cell, value in updates.items():
                ws[cell] = value  # Прямое присвоение значения по адресу ячейки
                ws[cell].alignment = Alignment(horizontal='center')  # Центрируем текст
                ws[cell].font = Font(name='Times New Roman', size=8)  # Устанавливаем шрифт


            wb.save(self.output_path)  # Сохраняем изменения
            wb.close()
        except Exception as e:
            print(f"Ошибка при обновлении файла: {e}")
            raise


class TemplateExcel(ExcelCopy):
    """
    Класс для работы с конкретным шаблоном Excel, наследующийся от ExcelCopy.
    """

    def __init__(self, template_path, output_path):
        super().__init__(template_path, output_path)

    def add_data(self, updates):
        """Метод для добавления данных в шаблон Excel."""
        self.create_copy()  # Создаем копию шаблона
        self.update_excel(updates)  # Обновляем файл с новыми данными

# Пример использования:
# updates = {
#             'AH12': today_date,
#             'AB18': 'Подотчетное лицо',
#             'BG18': 'Табельный',
#             'W21': 'Должность',
#             'BA21': 'Назначение аванса',
#             'AB23': 'Командирован в',
#             'X27': 'Название проекта ЦФО',
#             'P30': 'Срок командировки',
#             'AB36': 'Получено на карту',
#             'AB37': 'Авиабилеты',
#             'AB38': 'Ж/д билеты',
#             'AB39': '=SUM(AB36, P30, AB38)',  # Итоговая сумма
#             'AB40': 'Израсходовано',
#             'AB41': 'Остаток',  # Можно вычислить формулу в дальнейшем
#             'AB42': '=IF(AB41 < 0, AB40, 0)',
#             'AU9': '=AB39'
#             'AP14': day,  # Добавляем сегодняшнее число
#             'AW14': month,  # Добавляем сегодняшний месяц
#             'BE14': year,  # Добавляем сегодняшний год
#         }

# Список названий месяцев на русском
months = [
    "январь", "февраль", "март", "апрель", "май", "июнь",
    "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь"
]

if __name__ == '__main__':
    template_path = '../data/advance report template 2024.xlsx'  # Путь к шаблону
    output_path = '../data/newdata/new_advance report template 2024.xlsx'  # Путь для сохранения нового Excel-файла

    # Создаем экземпляр TemplateExcel
    template_excel = TemplateExcel(template_path, output_path)
    try:
        # Получаем сегодняшнюю дату в формате 'дд.мм.гггг'
        today_date = datetime.now()
        day = today_date.day
        month = months[today_date.month - 1]
        year = today_date.year

        # Обновляем значения в ячейках
        updates = {
            'AH12': today_date,
            'AP14': day,  # Добавляем сегодняшнее число
            'AW14': month,  # Добавляем сегодняшний месяц
            'BE14': year,  # Добавляем сегодняшний год
            'AB18': 'Иванов Иван Иваныч',
            'BG18': '000555',
            'W21': 'Ведущий Инженер ОТК',
            'BA21': 'Возврат средств',
            'AB23': 'г. Уфа',
            'X27': 'ЦФО',
            'P30': '25',
            'AB36': '31500',
            'AB37': '19900',
            'AB38': '0',
            'AB39': '50000,55',
            'AB40': '35000',
            'AB41': '15000,55',  # Можно вычислить формулу в дальнейшем
            'AB42': '0',
            'AU9': '=AB39',
            'X50': 'Санкт-Петербург - Уфа',
            'X51': 'Уфа - Санкт - Петербург',
            'X52': 'Такси',
            'X53': 'много много много много много слов',
        }

        # Добавляем данные в Excel
        template_excel.add_data(updates)  # Добавляем данные
    except Exception as e:
        print(f"Произошла ошибка: {e}")





