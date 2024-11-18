import pytest
import os
import openpyxl
from datetime import datetime
from src.example import TemplateExcel  # Замените на фактический путь, если нужно


@pytest.fixture
def setup_files(tmp_path):
    """Фикстура для создания временного шаблона и выходного файла."""
    # Путь к временным файлам
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"

    # Создаем временный Excel-файл как шаблон
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Тестовый шаблон"
    wb.save(template_path)

    yield str(template_path), str(output_path)

    # Удаляем файлы после теста (в pytest это делается автоматически)


def test_create_copy(setup_files):
    """Тест на создание копии шаблона Excel."""
    template_path, output_path = setup_files
    excel = TemplateExcel(template_path, output_path)

    excel.create_copy()

    # Проверяем, что файл скопирован
    assert os.path.exists(output_path), "Файл не был скопирован"


def test_update_excel(setup_files):
    """Тест на обновление значений в существующем Excel файле."""
    template_path, output_path = setup_files
    excel = TemplateExcel(template_path, output_path)
    excel.create_copy()

    # Указываем данные для обновления
    today_date = datetime.now().strftime('%d.%m.%Y')
    updates = {
        'A2': today_date,
        'A3': 'Тестовая запись',
        'B5': 12345,
    }

    # Обновляем Excel файл
    excel.update_excel(updates)

    # Проверяем, что данные записаны корректно
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    assert ws['A2'].value == today_date, "Дата не записана в ячейку A2"
    assert ws['A3'].value == 'Тестовая запись', "Тестовая запись не была добавлена в ячейку A3"
    assert ws['B5'].value == 12345, "Число 12345 не записано в ячейку B5"
    wb.close()


def test_add_data_method(setup_files):
    """Тест метода add_data в классе TemplateExcel."""
    template_path, output_path = setup_files
    excel = TemplateExcel(template_path, output_path)

    # Указываем данные для добавления
    today_date = datetime.now().strftime('%d.%m.%Y')
    updates = {
        'A2': today_date,
        'A3': 'Проверка метода add_data',
        'B5': 67890,
    }

    # Вызываем метод add_data, который должен сначала создать копию, затем обновить данные
    excel.add_data(updates)

    # Проверяем, что данные обновлены в копии файла
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    assert ws['A2'].value == today_date, "Дата не записана в ячейку A2 через add_data"
    assert ws['A3'].value == 'Проверка метода add_data', "Запись через add_data не была добавлена в ячейку A3"
    assert ws['B5'].value == 67890, "Число 67890 не записано в ячейку B5 через add_data"
    wb.close()
