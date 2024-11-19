import pytest
import os
from openpyxl import load_workbook
from src.work_with_excel import ExcelTemplate


@pytest.fixture
def setup_paths(tmp_path):
    """Фикстура для создания временных файлов."""
    # Пути к файлам
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"

    # Создаем временный Excel файл для шаблона
    wb = load_workbook(filename=template_path)
    ws = wb.active
    ws.title = "Шаблон"
    ws["A1"] = "Тестовые данные"
    wb.save(template_path)

    yield str(template_path), str(output_path)

    # Файлы удаляются после тестов (pytest делает это автоматически)


def test_write_new_file(setup_paths):
    """Тест создания нового файла на основе шаблона Excel."""
    template_path, output_path = setup_paths
    excel_template = ExcelTemplate(template_path, output_path)

    # Создаем новый файл
    excel_template.write(output_path)

    # Проверяем, что файл был создан
    assert os.path.exists(output_path), "Файл не был создан на основе шаблона"


def test_data_addition(setup_paths):
    """Тест добавления данных в Excel файл."""
    template_path, output_path = setup_paths
    excel_template = ExcelTemplate(template_path, output_path)

    # Создаем новый файл с добавленными данными
    excel_template.write(output_path)

    # Проверяем добавление данных в ячейки
    wb = load_workbook(output_path)
    ws = wb["Шаблон"]
    assert ws["G4"].value == "Вставляем данные", "Данные не были добавлены в ячейку G4"
    assert ws["G6"].value == "Еще данные", "Данные не были добавлены в ячейку G6"
    wb.close()


def test_configuration(setup_paths):
    """Тест конфигурации Excel файла (ширина и высота строк и столбцов)."""
    template_path, output_path = setup_paths
    excel_template = ExcelTemplate(template_path, output_path)

    # Создаем новый файл с настройками конфигурации
    excel_template.write(output_path)

    # Проверяем настройки конфигурации
    wb = load_workbook(output_path)
    ws = wb["Шаблон"]
    assert (
        ws.column_dimensions["A"].width == 30
    ), "Ширина столбца A не установлена правильно"
    assert (
        ws.row_dimensions[1].height == 30
    ), "Высота первой строки не установлена правильно"
    assert ws["C10"].value.startswith("="), "Формула не была добавлена в ячейку C10"
    wb.close()


def test_copy_formatting(setup_paths):
    """Тест копирования форматов (стилей) в новый файл."""
    template_path, output_path = setup_paths
    excel_template = ExcelTemplate(template_path, output_path)

    # Создаем новый файл с копированием форматов
    excel_template.write(output_path)

    # Проверяем, что форматы были скопированы
    wb = load_workbook(output_path)
    ws = wb["Шаблон"]
    assert ws["A1"].font.bold, "Шрифт в ячейке A1 не был скопирован как жирный"
    assert (
        ws["C3"].number_format == "$#,##0"
    ), "Формат валюты для столбца C не был установлен"
    wb.close()
